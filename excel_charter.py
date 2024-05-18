import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

#This function is used to read the excel file and separate columns into their own array abojects
def array_list(arr, main_sheet, modulus_value,rows,columns ):
    
    #each column in placed their respective arrays, aka score columns will be consolidated into one array, severity findings to a different one
    for i in range (1, rows + 1):
        for h in range (1, columns + 1) : 

            #modulus operator is used to determine if we are in the 1st, 2nd, or 3rd columns. The value is then appended to proper column
            if ( h % 3 == 1 and modulus_value == 1):
                cell_cve = main_sheet.cell(row=i, column=h)
                arr.append(cell_cve.value)
                
            if ( h % 3 == 2 and modulus_value == 2):
                base_score = main_sheet.cell(row=i, column=h)
                arr.append(base_score.value)
                
            if ( h % 3 == 0 and modulus_value == 3):
                severity_level = main_sheet.cell(row=i, column=h)
                arr.append(severity_level.value)
             
    return arr

#This function is used to create a pie chart for the severity findings 
def create_piechart(title,arr):
    
    #Creating an array that holds the classifications of severities
    severity_classification = ['Critical','High','Medium','Low']
    
    #Colors for the pie chart
    colors = ['red','blue','green','orange']
    
    #Initializing the severity count arrays
    severity_count =[0,0,0,0]

    #For loop used to iterate through the severity array and count each finding and update the corresponding count array
    for i in range (len(arr)):
        
        #Case statment used instead of elif statements to appropriately increment the correct position in the array 
        match arr[i]:
            case 'Critical':
                severity_count [0] += 1   
            case 'High':
                severity_count [1] += 1
            case 'Medium':
                severity_count [2] += 1
            case 'Low':
                severity_count [3] += 1
            case _:
                continue
    
    #Set up the Datafram for the chart info
    pie_info = pd.DataFrame ({ 'Name': severity_classification ,title: severity_count})
    
    #Create and show the pie chart
    pie_info.groupby(['Name']).sum().plot(kind='pie',y=title,autopct='%1.0f%%',colors=colors)
    plt.show()
    
    return 0


def create_boxplot(title,score_arr):
    
    colors = ['red']
    plt.boxplot(score_arr,patch_artist=True,boxprops=dict(facecolor='red',color='black'),whiskerprops=dict(color='black'),medianprops=dict(color='yellow'))
    plt.show()
    
    return

#define the main function
def main():

    #File name
    excel_path = 'CVE_File.xlsx'

    #Open the workbook
    workbook = openpyxl.load_workbook(excel_path)

    #Make the sheet active
    main_sheet = workbook.active

    #Maxmium rows in the excel sheet
    rows = main_sheet.max_row

    #Maximum columns in the excel sheet
    columns = main_sheet.max_column

    cve = []
    base_score = []
    finding = []

    #Call the array_list function to categorize the excel columns into different arrays
    cve = array_list(cve, main_sheet,1,rows,columns)
    base_score = array_list(base_score,main_sheet,2,rows,columns)
    finding = array_list(finding,main_sheet,3,rows,columns)
    
    #Call the create piechart function to create a piechart based on finding functions
    create_piechart("Severity", finding)

    #Call the create boxplot function to create a boxplot based on the base score
    create_boxplot("Base_Scores",base_score)
    
#Call the main function            
main()

    